"""
exporter.py
-----------
Exports accepted jobs to a formatted Excel workbook.
Produces three sheets: job listings, per-keyword summary, and run metadata.
"""

from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from workable_scraper.config import WEIGHTS
from workable_scraper.models import Job
from workable_scraper.profile import UserProfile

# ── Style constants ─────────────────────────────────────────────���─────────────

_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_ACCENT_FILL  = PatternFill("solid", fgColor="E8F4FD")
_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_WRAP   = Alignment(vertical="center", wrap_text=True)

_GREEN  = (PatternFill("solid", fgColor="C6EFCE"), Font(bold=True, color="276221"))
_YELLOW = (PatternFill("solid", fgColor="FFEB9C"), Font(bold=True, color="9C6500"))
_RED    = (PatternFill("solid", fgColor="FFC7CE"), Font(bold=True, color="9C0006"))

_COLUMNS = [
    ("Keyword",           20),
    ("Title",             35),
    ("Company",           25),
    ("Location",          20),
    ("Salary",            25),
    ("Exp. Required",     15),
    ("Remote",            30),
    ("Final Score",       12),
    ("Profile (30%)",     16),
    ("Salary (25%)",      16),
    ("Experience (20%)",  16),
    ("Remote (15%)",      16),
    ("Company (10%)",     16),
    ("Link",              55),
]

_SCORE_COLS = {8, 9, 10, 11, 12, 13}
_LINK_COL   = 14


# ── Public API ────────────────────────────���───────────────────────────────────

def export_to_excel(all_jobs: list[Job], filename: str, profile: UserProfile) -> str:
    """Writes all_jobs to an Excel workbook and returns the filename."""
    wb = openpyxl.Workbook()
    _write_jobs_sheet(wb, all_jobs)
    _write_summary_sheet(wb, all_jobs, profile)
    _write_info_sheet(wb, all_jobs, profile)
    wb.save(filename)
    print(f"\n  ✅ Excel saved: {filename}")
    return filename


# ── Sheet writers ───────────────────────────��─────────────────────────────────

def _write_jobs_sheet(wb: openpyxl.Workbook, jobs: list[Job]) -> None:
    ws = wb.active
    ws.title = "Jobs"

    for col, (header, width) in enumerate(_COLUMNS, 1):
        cell            = ws.cell(row=1, column=col, value=header)
        cell.font       = _HEADER_FONT
        cell.fill       = _HEADER_FILL
        cell.alignment  = _CENTER
        cell.border     = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 35

    for row_idx, job in enumerate(
        sorted(jobs, key=lambda j: j.score_final, reverse=True), 2
    ):
        fill = _ACCENT_FILL if row_idx % 2 == 0 else PatternFill()

        salary_display = job.salary_raw
        if job.salary_min_usd:
            if job.salary_max_usd and job.salary_max_usd != job.salary_min_usd:
                salary_display = (
                    f"${job.salary_min_usd:,.0f} - ${job.salary_max_usd:,.0f} USD/yr"
                    f"\n({job.salary_raw})"
                )
            else:
                salary_display = f"${job.salary_min_usd:,.0f} USD/yr\n({job.salary_raw})"

        exp_display = (
            f"{job.experience_required:.0f} yrs"
            if job.experience_required is not None else "Not listed"
        )

        row_data = [
            job.keyword,
            job.title,
            job.company,
            job.location or "Not specified",
            salary_display,
            exp_display,
            job.remote_info,
            job.score_final,
            job.score_profile,
            job.score_salary,
            job.score_experience,
            job.score_remote,
            job.score_company,
            job.url,
        ]

        for col, value in enumerate(row_data, 1):
            cell           = ws.cell(row=row_idx, column=col, value=value)
            cell.border    = _BORDER
            cell.alignment = _CENTER if col in _SCORE_COLS else _WRAP
            if fill.fgColor.rgb != "00000000":
                cell.fill = fill

        score_cell = ws.cell(row=row_idx, column=8)
        if job.score_final >= 8.0:
            score_cell.fill, score_cell.font = _GREEN
        elif job.score_final >= 6.5:
            score_cell.fill, score_cell.font = _YELLOW
        else:
            score_cell.fill, score_cell.font = _RED

        link_cell           = ws.cell(row=row_idx, column=_LINK_COL)
        link_cell.hyperlink = job.url
        link_cell.font      = Font(color="0563C1", underline="single")

        ws.row_dimensions[row_idx].height = 40


def _write_summary_sheet(
    wb: openpyxl.Workbook,
    jobs: list[Job],
    profile: UserProfile,
) -> None:
    ws = wb.create_sheet("Summary")
    cols = [("Keyword", 30), ("Jobs Accepted", 15), ("Avg Score", 15), ("Best Score", 20)]
    for col, (header, width) in enumerate(cols, 1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border    = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    by_keyword: dict[str, list[Job]] = defaultdict(list)
    for job in jobs:
        by_keyword[job.keyword].append(job)

    for row_idx, kw in enumerate(profile.search_keywords, 2):
        kw_jobs   = by_keyword.get(kw, [])
        avg_score = round(sum(j.score_final for j in kw_jobs) / len(kw_jobs), 2) if kw_jobs else 0
        best      = round(max((j.score_final for j in kw_jobs), default=0), 2)
        for col, val in enumerate([kw, len(kw_jobs), avg_score, best], 1):
            cell           = ws.cell(row=row_idx, column=col, value=val)
            cell.border    = _BORDER
            cell.alignment = _CENTER


def _write_info_sheet(
    wb: openpyxl.Workbook,
    jobs: list[Job],
    profile: UserProfile,
) -> None:
    ws = wb.create_sheet("Info")
    ws.column_dimensions["A"].width = 70

    strict_note = ""
    if profile.strict_experience:
        max_allowed = profile.experience_years + profile.experience_gap
        strict_note = (
            f"  |  Gap: {profile.experience_gap} yrs  "
            f"|  Max allowed: {max_allowed:.1f} yrs"
        )

    rows = [
        ("Workable Job Scraper — Run Info", True),
        (f"Generated:               {datetime.now().strftime('%Y-%m-%d %H:%M')}", False),
        (f"User:                    {profile.name}", False),
        (f"Total jobs accepted:     {len(jobs)}", False),
        ("", False),
        (f"Experience:              {profile.experience_years} yrs", False),
        (f"Strict experience filter: {'ON' if profile.strict_experience else 'OFF'}{strict_note}", False),
        (f"Salary target:           ${profile.salary_target_usd:,}/yr (USD)", False),
        (f"Min score:               {profile.min_score}", False),
        ("", False),
        (f"Keywords searched:       {', '.join(profile.search_keywords)}", False),
        ("", False),
        (
            f"Weights — profile: {WEIGHTS['profile']:.0%}  "
            f"salary: {WEIGHTS['salary']:.0%}  "
            f"experience: {WEIGHTS['experience']:.0%}  "
            f"remote: {WEIGHTS['remote']:.0%}  "
            f"company: {WEIGHTS['company']:.0%}",
            False,
        ),
    ]

    for row_idx, (text, bold) in enumerate(rows, 1):
        cell      = ws.cell(row=row_idx, column=1, value=text)
        cell.font = Font(bold=bold, size=13 if bold else 11)
